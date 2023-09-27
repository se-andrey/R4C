from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime, timedelta

from django.utils import timezone

from robots.models import Robot


def summary_excel():
    # Данные из базы за последнюю неделю
    end_date = datetime.now(timezone.utc)
    start_date = end_date - timedelta(days=7)
    robots = Robot.objects.filter(created__gte=start_date, created__lte=end_date)

    # Создаем книгу Excel
    workbook = Workbook()

    # Проверяем найдены ли записи в БД
    if not robots:
        sheet = workbook.create_sheet(
            title=f'{start_date.date()}-{end_date.date()}'
        )
        sheet.append([f'За период с {start_date.date()} по {end_date.date()} ничего не произведено'])

    else:

        # Словарь для хранения данных по роботам
        model_data = {}
        for robot in robots:
            model = robot.model
            version = robot.version
            if model not in model_data:
                model_data[model] = {}
            if version not in model_data[model]:
                model_data[model][version] = 1
            else:
                model_data[model][version] += 1

        for model, version_data in model_data.items():

            # Создаем новый лист для модели
            sheet = workbook.create_sheet(
                title=f'{model}_{start_date.date()}-{end_date.date()}'
            )
            sheet.append(['Модель', 'Версия', 'Количество за неделю'])
            sheet['A1'].font = Font(bold=True)
            sheet['B1'].font = Font(bold=True)
            sheet['C1'].font = Font(bold=True)

            # Заполняем страницу с версиями модели и количеством произведенных штук
            for version, count in version_data.items():
                sheet.append([model, version, count])

    workbook.remove(workbook.get_sheet_by_name('Sheet'))

    return workbook
